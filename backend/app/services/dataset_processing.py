"""Dataset file parsing and metadata extraction.

Pure parsing: no database, no HTTP. The service layer calls
:func:`extract_metadata` and decides what to persist.

Memory discipline: row counts come from streaming/chunked reads and column
types are inferred from a bounded sample, so a large upload never has to be
materialised in full.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, BinaryIO

import pandas as pd

from app.core.config import settings
from app.core.exceptions import ProcessingError
from app.core.logging import get_logger
from app.models.dataset import DatasetFileType

logger = get_logger(__name__)

#: Encodings tried in order for CSV files.
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")


@dataclass
class ColumnInfo:
    name: str
    dtype: str
    nullable: bool = False


@dataclass
class DatasetMetadata:
    row_count: int
    column_count: int
    columns: list[ColumnInfo] = field(default_factory=list)

    def columns_as_dicts(self) -> list[dict[str, Any]]:
        return [
            {"name": c.name, "dtype": c.dtype, "nullable": c.nullable} for c in self.columns
        ]


def _normalise_dtype(series: pd.Series) -> str:
    """Map a pandas dtype onto a small, stable vocabulary."""
    if pd.api.types.is_bool_dtype(series):
        return "boolean"
    if pd.api.types.is_integer_dtype(series):
        return "integer"
    if pd.api.types.is_float_dtype(series):
        return "float"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "datetime"
    return "string"


def _columns_from_frame(sample: pd.DataFrame) -> list[ColumnInfo]:
    return [
        ColumnInfo(
            name=str(name),
            dtype=_normalise_dtype(sample[name]),
            nullable=bool(sample[name].isna().any()),
        )
        for name in sample.columns
    ]


def _detect_csv_encoding(path: Path) -> str:
    """Return the first encoding that can decode the opening block."""
    for encoding in CSV_ENCODINGS:
        try:
            with path.open("r", encoding=encoding) as handle:
                handle.read(64 * 1024)
            return encoding
        except (UnicodeDecodeError, LookupError):
            continue
    # latin-1 decodes any byte sequence, so this is a genuine read failure.
    raise ProcessingError("The file's text encoding could not be determined.")


def _detect_csv_delimiter(path: Path, encoding: str) -> str:
    """Sniff the delimiter, falling back to a comma."""
    try:
        with path.open("r", encoding=encoding, newline="") as handle:
            sample = handle.read(32 * 1024)
        if not sample.strip():
            raise ProcessingError("The file is empty.")
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def extract_csv_metadata(path: Path) -> DatasetMetadata:
    """Parse a CSV file, counting rows in chunks rather than loading it whole."""
    encoding = _detect_csv_encoding(path)
    delimiter = _detect_csv_delimiter(path, encoding)

    try:
        sample = pd.read_csv(
            path,
            encoding=encoding,
            sep=delimiter,
            nrows=settings.DATASET_TYPE_SAMPLE_ROWS,
        )
    except pd.errors.EmptyDataError as exc:
        raise ProcessingError("The file contains no data.") from exc
    except (pd.errors.ParserError, ValueError) as exc:
        raise ProcessingError("The CSV file is malformed and could not be parsed.") from exc

    if sample.columns.empty:
        raise ProcessingError("No columns were found in the file.")

    # Count every row without holding more than one chunk at a time.
    row_count = 0
    try:
        for chunk in pd.read_csv(
            path,
            encoding=encoding,
            sep=delimiter,
            chunksize=settings.DATASET_CSV_CHUNK_ROWS,
            usecols=[0],
        ):
            row_count += len(chunk)
    except (pd.errors.ParserError, ValueError) as exc:
        raise ProcessingError("The CSV file is malformed and could not be parsed.") from exc

    return DatasetMetadata(
        row_count=row_count,
        column_count=len(sample.columns),
        columns=_columns_from_frame(sample),
    )


def extract_xlsx_metadata(path: Path) -> DatasetMetadata:
    """Parse the first worksheet of an .xlsx workbook."""
    try:
        from openpyxl import load_workbook

        # read_only streams rows instead of building the whole sheet in memory.
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a variety of low-level errors
        raise ProcessingError("The Excel file is malformed and could not be read.") from exc

    try:
        worksheet = workbook.worksheets[0] if workbook.worksheets else None
        if worksheet is None:
            raise ProcessingError("The workbook contains no worksheets.")

        # max_row includes the header; it can be None for a truly empty sheet.
        total_rows = worksheet.max_row or 0
        row_count = max(total_rows - 1, 0)
    finally:
        workbook.close()

    try:
        sample = pd.read_excel(
            path,
            sheet_name=0,
            nrows=settings.DATASET_TYPE_SAMPLE_ROWS,
            engine="openpyxl",
        )
    except ValueError as exc:
        raise ProcessingError("The Excel file could not be parsed.") from exc

    if sample.columns.empty:
        raise ProcessingError("No columns were found in the first worksheet.")

    return DatasetMetadata(
        row_count=row_count,
        column_count=len(sample.columns),
        columns=_columns_from_frame(sample),
    )


def extract_metadata(path: Path, file_type: DatasetFileType) -> DatasetMetadata:
    """Dispatch to the parser for ``file_type``."""
    if file_type is DatasetFileType.CSV:
        return extract_csv_metadata(path)
    if file_type is DatasetFileType.XLSX:
        return extract_xlsx_metadata(path)
    raise ProcessingError("Unsupported file type.")


def looks_empty(source: BinaryIO) -> bool:
    """True when the incoming stream holds no bytes. Rewinds the stream."""
    position = source.tell()
    is_empty = not source.read(1)
    source.seek(position)
    return is_empty
