"""Report builder and renderer tests.

Two things are worth guarding here. First, that section availability is driven
by the dataset's detected roles rather than by column names - a dataset with no
dates must not produce a trend section, and must say why. Second, that all four
renderers consume the same :class:`ReportData` and produce a well-formed file,
including for the awkward cases (no sections at all, a very wide table).
"""

from __future__ import annotations

import csv
import io
import uuid
import zipfile

import pandas as pd
import pytest

from app.models.dataset import Dataset, DatasetFileType, DatasetStatus
from app.models.report import ReportFileFormat, ReportTemplateName
from app.schemas.report import (
    ReportData,
    ReportSection,
    ReportSectionKey,
    ReportTable,
)
from app.services import (
    ai_analyst_service,
    data_quality,
    dataset_access,
    dataset_profiling,
    insights_service,
    report_builder,
    report_csv,
    report_pdf,
    report_pptx,
    report_service,
    report_xlsx,
    semantic_columns,
)

ROWS = 300
Key = ReportSectionKey


@pytest.fixture
def sales_frame() -> pd.DataFrame:
    """A dataset rich enough to support every section."""
    return pd.DataFrame(
        {
            "order_id": range(1, ROWS + 1),
            "customer_id": [1000 + (index % 40) for index in range(ROWS)],
            "revenue": [round(10.5 + (index % 97) * 1.37, 2) for index in range(ROWS)],
            "quantity": [(index % 5) + 1 for index in range(ROWS)],
            "category": [["Electronics", "Grocery", "Apparel"][index % 3] for index in range(ROWS)],
            "order_date": pd.date_range("2024-01-01", periods=ROWS, freq="D"),
        }
    )


@pytest.fixture
def bare_frame() -> pd.DataFrame:
    """No dates, no identifiers, one measure - most sections are impossible."""
    return pd.DataFrame(
        {
            "label": [f"item-{index % 7}" for index in range(50)],
            "score": [float(index % 11) for index in range(50)],
        }
    )


def make_loaded(frame: pd.DataFrame, name: str = "Sales export") -> dataset_access.LoadedDataset:
    dataset = Dataset(
        id=uuid.uuid4(),
        name=name,
        original_filename=f"{name}.csv",
        storage_key="datasets/test/data.csv",
        file_type=DatasetFileType.CSV,
        file_size=1,
        status=DatasetStatus.READY,
        project_id=uuid.uuid4(),
    )
    return dataset_access.LoadedDataset(dataset=dataset, frame=frame, version=None)


def build(
    frame: pd.DataFrame,
    template: ReportTemplateName = ReportTemplateName.FULL,
    sections: list[ReportSectionKey] | None = None,
    *,
    with_insights: bool = True,
) -> ReportData:
    """The same path ``report_service`` takes, without database or storage."""
    loaded = make_loaded(frame)
    profile = dataset_profiling.profile_frame(frame, dataset_id=loaded.dataset.id)
    quality = data_quality.assess_quality(profile, frame, dataset_id=loaded.dataset.id)
    analyst = ai_analyst_service.build_report(frame, loaded, profile=profile, quality=quality)
    insights = (
        insights_service.build_deterministic(
            frame,
            loaded,
            analyst,
            quality,
            project_id=loaded.dataset.project_id,
            generated_by="Test User",
        )
        if with_insights
        else None
    )
    return report_builder.build(
        frame,
        loaded,
        analyst,
        profile,
        quality,
        project_id=loaded.dataset.project_id,
        template=template,
        sections=sections,
        title=None,
        generated_by="Test User",
        insights=insights,
    )


def section(data: ReportData, key: ReportSectionKey) -> object | None:
    return next((item for item in data.sections if item.key is key), None)


# --- Availability ------------------------------------------------------------


def test_rich_dataset_supports_every_section(sales_frame: pd.DataFrame) -> None:
    model = semantic_columns.detect(sales_frame)
    supported, reasons = report_builder.available_sections(model)

    assert not reasons, f"unexpectedly unavailable: {reasons}"
    assert set(supported) == set(report_builder.SECTION_ORDER)


def test_bare_dataset_reports_why_sections_are_unavailable(bare_frame: pd.DataFrame) -> None:
    model = semantic_columns.detect(bare_frame)
    supported, reasons = report_builder.available_sections(model)

    # No date column, so nothing time-based can be produced.
    assert Key.TRENDS not in supported
    assert Key.FORECAST not in supported
    assert Key.COHORT not in supported
    # A single measure is not enough to cluster or correlate.
    assert Key.SEGMENTATION not in supported
    assert Key.CORRELATION not in supported
    # These need nothing in particular and must always be offered.
    for key in (Key.EXECUTIVE_SUMMARY, Key.DATASET_OVERVIEW, Key.DATA_QUALITY, Key.KPIS):
        assert key in supported

    assert all(reason and reason.strip() for reason in reasons.values())


def test_unsupported_sections_are_skipped_with_a_reason(bare_frame: pd.DataFrame) -> None:
    data = build(bare_frame)

    assert section(data, Key.TRENDS) is None
    skipped = {item["section"]: item["reason"] for item in data.skipped}
    assert Key.TRENDS.value in skipped
    assert "date" in skipped[Key.TRENDS.value].lower()


def test_sections_follow_canonical_order_not_request_order(sales_frame: pd.DataFrame) -> None:
    data = build(sales_frame, sections=[Key.RECOMMENDATIONS, Key.KPIS, Key.EXECUTIVE_SUMMARY])
    keys = [item.key for item in data.sections]

    assert keys == [Key.EXECUTIVE_SUMMARY, Key.KPIS, Key.RECOMMENDATIONS]


def test_templates_select_different_sections(sales_frame: pd.DataFrame) -> None:
    executive = {item.key for item in build(sales_frame, ReportTemplateName.EXECUTIVE).sections}
    customer = {item.key for item in build(sales_frame, ReportTemplateName.CUSTOMER).sections}

    assert Key.RFM in customer and Key.RFM not in executive
    assert Key.DATASET_OVERVIEW in executive


# --- Content -----------------------------------------------------------------


def test_report_reuses_the_analyst_figures(sales_frame: pd.DataFrame) -> None:
    """KPIs must be the analyst's, not a second computation."""
    loaded = make_loaded(sales_frame)
    analyst = ai_analyst_service.build_report(sales_frame, loaded)
    data = build(sales_frame, sections=[Key.KPIS])

    kpi_section = section(data, Key.KPIS)
    assert kpi_section is not None
    reported = {row[0]: row[3] for row in kpi_section.tables[0].rows}  # type: ignore[attr-defined]
    for kpi in analyst.kpis:
        if kpi.available and kpi.name in reported:
            assert reported[kpi.name] == kpi.value


def test_tables_are_truncated_with_a_visible_note(sales_frame: pd.DataFrame) -> None:
    data = build(sales_frame, sections=[Key.DATASET_OVERVIEW, Key.RFM])
    rfm = section(data, Key.RFM)
    assert rfm is not None

    customers = rfm.tables[1]  # type: ignore[attr-defined]
    assert len(customers.rows) <= report_builder.TABLE_ROW_LIMIT
    if customers.note:
        assert "Showing the first" in customers.note


def test_no_ai_provider_still_produces_a_narrative(sales_frame: pd.DataFrame) -> None:
    """The deterministic summary is the source of truth; AI is additive."""
    data = build(sales_frame, sections=[Key.EXECUTIVE_SUMMARY, Key.AI_INSIGHTS])
    summary = section(data, Key.EXECUTIVE_SUMMARY)

    assert summary is not None
    assert summary.narrative and summary.narrative[0].strip()  # type: ignore[attr-defined]
    # With no provider configured the AI section is skipped, not faked.
    assert section(data, Key.AI_INSIGHTS) is None
    assert data.ai_available is False


def test_cell_text_formats_numbers_consistently() -> None:
    assert report_builder.cell_text(None) == "-"
    assert report_builder.cell_text(1234) == "1,234"
    assert report_builder.cell_text(1234.5) == "1,234.50"
    assert report_builder.cell_text(1234.0) == "1,234"
    assert report_builder.cell_text(float("nan")) == "-"
    assert report_builder.cell_text(True) == "yes"


# --- Renderers ---------------------------------------------------------------


@pytest.fixture
def rendered(sales_frame: pd.DataFrame) -> ReportData:
    return build(sales_frame)


def test_pdf_renders_a_valid_document(rendered: ReportData) -> None:
    payload = report_pdf.render(rendered)

    assert payload.startswith(b"%PDF-")
    assert payload.rstrip().endswith(b"%%EOF")


def test_xlsx_renders_one_sheet_per_section(rendered: ReportData) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(report_xlsx.render(rendered)))

    assert workbook.sheetnames[0] == "Report"
    # Cover sheet plus one sheet per rendered section.
    assert len(workbook.sheetnames) == len(rendered.sections) + 1
    assert len(set(workbook.sheetnames)) == len(workbook.sheetnames)


def test_xlsx_keeps_numbers_numeric(sales_frame: pd.DataFrame) -> None:
    from openpyxl import load_workbook

    data = build(sales_frame, sections=[Key.DATASET_OVERVIEW])
    workbook = load_workbook(io.BytesIO(report_xlsx.render(data)))
    sheet = workbook[workbook.sheetnames[1]]

    numeric = [
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
    ]
    assert numeric, "the overview sheet wrote every number as text"


def test_csv_labels_every_section(rendered: ReportData) -> None:
    payload = report_csv.render(rendered)

    assert payload.startswith(b"\xef\xbb\xbf"), "Excel needs the UTF-8 BOM"
    text = payload.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    markers = {row[0] for row in rows if row and row[0].startswith("## ")}

    for item in rendered.sections:
        assert f"## {item.title}" in markers


def test_pptx_renders_a_readable_deck(rendered: ReportData) -> None:
    payload = report_pptx.render(rendered)

    # A .pptx is a zip container; this proves it is well-formed.
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert "ppt/presentation.xml" in archive.namelist()
        assert archive.testzip() is None


def test_pptx_splits_long_tables_across_slides() -> None:
    from pptx import Presentation

    data = ReportData(
        title="Wide",
        project_id=uuid.uuid4(),
        dataset_id=uuid.uuid4(),
        dataset_name="Wide",
        version_label="Original dataset",
        template=ReportTemplateName.FULL,
        generated_at=pd.Timestamp("2026-01-01", tz="UTC").to_pydatetime(),
        generated_by="Test User",
        row_count=1,
        column_count=1,
        sections=[
            ReportSection(
                key=Key.KPIS,
                title="Key performance indicators",
                tables=[
                    ReportTable(
                        columns=["a", "b"],
                        rows=[[index, index * 2] for index in range(25)],
                    )
                ],
            )
        ],
    )
    deck = Presentation(io.BytesIO(report_pptx.render(data)))

    # 25 rows at 10 per slide is three table slides, plus title and agenda.
    assert len(deck.slides) == 5


def test_every_renderer_handles_a_report_with_no_sections() -> None:
    data = ReportData(
        title="Empty",
        project_id=uuid.uuid4(),
        dataset_id=uuid.uuid4(),
        dataset_name="Empty",
        version_label="Original dataset",
        template=ReportTemplateName.EXECUTIVE,
        generated_at=pd.Timestamp("2026-01-01", tz="UTC").to_pydatetime(),
        generated_by="Test User",
        row_count=0,
        column_count=0,
        skipped=[{"section": "trends", "reason": "No date column."}],
    )

    assert report_pdf.render(data).startswith(b"%PDF-")
    assert report_xlsx.render(data)
    assert b"trends" in report_csv.render(data)
    assert report_pptx.render(data)


# --- Storage keys ------------------------------------------------------------


@pytest.mark.parametrize(
    "name",
    ["../../etc/passwd", "..", "  ", "report/../secret", "\\windows\\system32", "🙂"],
)
def test_storage_keys_cannot_escape_the_reports_prefix(name: str) -> None:
    report_id = uuid.uuid4()
    key = report_service.storage_key_for(report_id, name, ReportFileFormat.PDF)

    assert key.startswith(f"reports/{report_id}/")
    assert ".." not in key
    assert key.count("/") == 2
    assert key.endswith(".pdf")


def test_slugify_never_returns_an_empty_name() -> None:
    assert report_service.slugify("") == "report"
    assert report_service.slugify("!!!") == "report"
    assert report_service.slugify("Q3 Sales Review") == "q3-sales-review"


# --- AI Insights integration -------------------------------------------------


def test_insight_sections_reuse_the_insight_engine(sales_frame: pd.DataFrame) -> None:
    """Business health, opportunities and risks come from the same report."""
    data = build(
        sales_frame,
        sections=[Key.BUSINESS_HEALTH, Key.CRITICAL_INSIGHTS, Key.OPPORTUNITIES, Key.RISKS],
    )
    keys = [item.key for item in data.sections]

    assert Key.BUSINESS_HEALTH in keys
    health = section(data, Key.BUSINESS_HEALTH)
    assert health is not None
    # The methodology travels with the section, so a PDF can explain the score.
    assert any("weighted average" in text for text in health.narrative)  # type: ignore[attr-defined]


def test_insight_sections_are_skipped_when_the_engine_was_not_run(
    sales_frame: pd.DataFrame,
) -> None:
    data = build(sales_frame, sections=[Key.BUSINESS_HEALTH], with_insights=False)

    assert section(data, Key.BUSINESS_HEALTH) is None
    reasons = {item["section"]: item["reason"] for item in data.skipped}
    assert Key.BUSINESS_HEALTH.value in reasons
    assert "Insights engine" in reasons[Key.BUSINESS_HEALTH.value]


def test_recommendations_include_the_prioritised_plan(sales_frame: pd.DataFrame) -> None:
    data = build(sales_frame, sections=[Key.RECOMMENDATIONS])
    recommendations = section(data, Key.RECOMMENDATIONS)

    assert recommendations is not None
    titles = [table.title for table in recommendations.tables]  # type: ignore[attr-defined]
    assert "Prioritised actions" in titles


def test_every_renderer_handles_the_insight_sections(sales_frame: pd.DataFrame) -> None:
    data = build(sales_frame, ReportTemplateName.EXECUTIVE)

    assert report_pdf.render(data).startswith(b"%PDF-")
    assert report_xlsx.render(data)
    assert report_csv.render(data)
    assert report_pptx.render(data)
